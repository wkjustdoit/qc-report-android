# -*- coding: utf-8 -*-
"""
Excel 质检报告处理：填表、删旧图、按单元格从左到右平铺插入新图。
"""
import io
import os
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.utils import get_column_letter
from PIL import Image

from image_utils import compress_image


EMU_PER_PX = 9525   # 1 像素 = 9525 EMU
PT_PER_PX = 0.75    # Excel 行高（点）与像素换算：1px ≈ 0.75pt


def parse_cell(addr: str):
    """解析 'D6' -> (row=6, col=4)。"""
    addr = addr.strip().upper()
    col_str = ""
    row_str = ""
    for ch in addr:
        if ch.isalpha():
            col_str += ch
        else:
            row_str += ch
    col = 0
    for ch in col_str:
        col = col * 26 + (ord(ch) - ord('A') + 1)
    row = int(row_str)
    return row, col


def row_height_pt(ws, row: int):
    """获取行高（点）。未显式设置时返回默认 15。"""
    rd = ws.row_dimensions.get(row)
    if rd and rd.height:
        return float(rd.height)
    return 15.0


def _col_width_px(ws, col: int):
    """单列宽度（像素）。col 为 1-based 整数。"""
    letter = get_column_letter(col)
    cd = ws.column_dimensions.get(letter)
    if cd and cd.width:
        return cd.width * 7 + 5
    return 64  # Excel 默认列宽约 8.43 字符 ≈ 64px


def cell_width_px(ws, row: int, col: int):
    """
    目标单元格（含合并区域）的总宽度（像素）。
    例如 D107 合并为 D107:G107，则返回 D~G 四列宽度之和。
    """
    merge = None
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            merge = mr
            break
    if merge:
        return sum(_col_width_px(ws, c) for c in range(merge.min_col, merge.max_col + 1))
    return _col_width_px(ws, col)


def anchor_cell_addr(anchor):
    """把 openpyxl 图片 anchor 转成 'A1' 字符串（基于起始单元格）。"""
    if hasattr(anchor, '_from'):
        return "{}{}".format(
            get_column_letter(anchor._from.col + 1),
            anchor._from.row + 1
        )
    return ""


def delete_images_at_cell(ws, cell_addr: str):
    """删除锚点起始单元格等于 cell_addr 的所有图片。"""
    target = cell_addr.strip().upper()
    kept = []
    removed = 0
    for img in ws._images:
        if anchor_cell_addr(img.anchor).upper() == target:
            removed += 1
            continue
        kept.append(img)
    ws._images = kept
    return removed


def insert_photo_groups(ws, photo_groups: dict, image_config: dict):
    """
    把照片按组插入到指定单元格，组内多张照片从左到右依次平铺。

    防变形 / 防重叠 / 防少图 的核心做法：
      - 每张照片先用 PIL 精确缩放到“目标显示框”的像素尺寸（base_h 决定高度、原图
        宽高比决定宽度），再写成图片字节。这样“图片字节尺寸 == 显示框尺寸”，
        无论 WPS/Excel 按哪种方式渲染都不会拉伸变形。
      - 基准高度 = 目标行行高（像素）；行高不足抬到 min_height_px，使照片与单元格
        严丝合缝；行高更高的（如 D108/D109 的 135pt）保持满高不压低。
      - 组内总宽超过【单元格(含合并区)可用宽】时，整组按原比例等比缩小，保证全部
        落在单元格内、互不重叠；并预留安全边距，避免最右一张被裁掉（即“少一张”）。
      - 多张照片锚定同一行，水平偏移依次累加（带固定间隙）-> 从左到右排开。
    """
    gap = image_config.get("gap_px", 6)
    min_h = image_config.get("min_height_px", 110)        # 行高不足的组至少抬到这个高度
    margin = image_config.get("fit_margin_px", 8)         # 单元格左右内边距
    safety = image_config.get("fit_safety_px", 12)        # 额外安全边距，防最右一张被裁
    pad = image_config.get("pad_px", 2)                   # 照片上下留白（像素）
    quality = image_config.get("quality", 90)
    target_kb = image_config.get("target_size_kb", 700)
    max_dim = image_config.get("max_dimension", 2048)
    rotate = image_config.get("auto_rotate_portrait", True)

    summary = {}
    for cell_addr, paths in photo_groups.items():
        if not paths:
            continue

        delete_images_at_cell(ws, cell_addr)

        from_row, from_col = parse_cell(cell_addr)

        # 目标行高（pt）-> 像素；照片高度与之对齐，必要时抬行
        rd = ws.row_dimensions[from_row]
        row_pt = float(rd.height) if rd.height else 15.0
        row_px = row_pt / PT_PER_PX
        base_h = max(row_px, min_h)            # 基准显示高度

        # 单元格（含合并区）可用宽度，扣掉左右内边距 + 额外安全边距
        avail = max(cell_width_px(ws, from_row, from_col) - 2 * margin - safety, 40)

        # 打开 + 压缩到约 700KB（高分辨率，放大仍清晰）+ 记录处理后尺寸
        # 注意：图片字节【保持高分辨率】，只在 Excel 里“显示缩小”渲染，
        # 绝不把它缩到显示框物理尺寸再存 PNG（那会丢失高清信息，放大必然糊）。
        prepared = []
        for p in paths:
            data = compress_image(
                p, target_size_kb=target_kb, quality=quality,
                max_dimension=max_dim, auto_rotate_portrait=rotate,
            )
            tmp = Image.open(io.BytesIO(data))
            pw, ph = tmp.size
            base_w = pw * (base_h / max(ph, 1))
            prepared.append((data, pw, ph, base_w))

        sum_w = sum(bw for _, _, _, bw in prepared)
        total_needed = sum_w + gap * (len(prepared) - 1)

        # 放不下 -> 仅缩小【显示尺寸】（字节仍高分辨率），保证全部落在单元格内、不重叠
        eff_h = base_h
        if total_needed > avail:
            eff_h = base_h * (avail - gap * (len(prepared) - 1)) / max(sum_w, 1)

        # 仅当照片比行高时才抬行，使照片填满该行
        if eff_h > row_px + 0.5:
            rd.height = eff_h * PT_PER_PX

        # 目标列（含合并区）的像素宽列表，用于把绝对像素换算成 (col, colOff)。
        # 关键：TwoCellAnchor 的 to.col 必须指向“终点实际所在列”，不能一律用起始列
        # + 超大 colOff。Excel 桌面版会把 colOff 跨列累加渲染正常，但 WPS 会把
        # 「to.col=D 且 colOff 超出 D 列宽」的图片截断/拉伸到 D 列内，导致跨列图片
        # 变形、最右一张被裁掉（即“少一张”）。故这里显式换算到真实列，保证两软件一致。
        _mr = None
        for m in ws.merged_cells.ranges:
            if m.min_row <= from_row <= m.max_row and m.min_col <= from_col <= m.max_col:
                _mr = m
                break
        _cols = list(range(_mr.min_col, _mr.max_col + 1)) if _mr else [from_col]
        _col_w_px = [_col_width_px(ws, c) for c in _cols]

        def px_to_coloff(x_px):
            """把相对起始列左缘的绝对像素 x_px，换算成 (0-based列号, 列内colOff_EMU)。"""
            x = max(0.0, float(x_px))
            c = _cols[0] - 1
            acc = 0.0
            for w in _col_w_px:
                if x <= acc + w:
                    return c, int(round((x - acc) * EMU_PER_PX))
                acc += w
                c += 1
            # 超出合并区右缘：落在最后一列右侧（仍合法，渲染器按绝对坐标定位）
            return c, int(round((x - acc) * EMU_PER_PX))

        cum_x = 0.0
        inserted = 0
        for data, pw, ph, base_w in prepared:
            # 显示框：高度 eff_h，宽度按原图比例推算（不变形）
            disp_w = max(pw * (eff_h / max(ph, 1)), 1.0)
            disp_h = eff_h

            # TwoCellAnchor 锁定矩形（from 左上 + to 右下），且 to 落在其真实所在列，
            # WPS 严格按像素渲染不变形；字节保持高分辨率，缩小显示清晰。
            fcol, foff = px_to_coloff(cum_x)
            tcol, toff = px_to_coloff(cum_x + disp_w)
            xl_img = XLImage(io.BytesIO(data))
            xl_img.anchor = TwoCellAnchor(
                _from=AnchorMarker(
                    col=fcol,
                    colOff=foff,
                    row=from_row - 1,
                    rowOff=int(round(pad * EMU_PER_PX)),
                ),
                to=AnchorMarker(
                    col=tcol,
                    colOff=toff,
                    row=from_row - 1,
                    rowOff=int(round((pad + disp_h) * EMU_PER_PX)),
                ),
            )
            ws.add_image(xl_img)
            cum_x = cum_x + disp_w + gap
            inserted += 1

        summary[cell_addr] = {
            "found": len(paths),
            "inserted": inserted,
            "cells": [cell_addr],
        }
    return summary


def fill_fields(ws, fields: dict, packaging_info: dict,
                sample_cartons, sample_numbers, total_cartons):
    """把包装要求信息和抽检信息写入对应单元格。"""
    if packaging_info.get("quantity") is not None and fields.get("quantity_cell"):
        row, col = parse_cell(fields["quantity_cell"])
        ws.cell(row=row, column=col, value=packaging_info["quantity"])

    if packaging_info.get("order_no") and fields.get("order_no_cell"):
        row, col = parse_cell(fields["order_no_cell"])
        ws.cell(row=row, column=col, value=packaging_info["order_no"])

    if fields.get("date_cell"):
        row, col = parse_cell(fields["date_cell"])
        cell = ws.cell(row=row, column=col, value=datetime.now().date())
        cell.number_format = 'yyyy-mm-dd'

    if packaging_info.get("size") and fields.get("size_cell"):
        row, col = parse_cell(fields["size_cell"])
        ws.cell(row=row, column=col, value=packaging_info["size"])

    def write_num(addr, val):
        if addr and val is not None and str(val).strip() != "":
            row, col = parse_cell(addr)
            try:
                ws.cell(row=row, column=col, value=int(val))
            except ValueError:
                ws.cell(row=row, column=col, value=val)

    write_num(fields.get("sample_cartons_cell"), sample_cartons)
    write_num(fields.get("sample_numbers_cell"), sample_numbers)
    write_num(fields.get("total_cartons_cell"), total_cartons)


def process_excel(template_path: str, output_path: str, packaging_info: dict,
                  photo_groups: dict, fields: dict, image_config: dict,
                  sample_cartons=None, sample_numbers=None, total_cartons=None):
    """
    完整处理流程：打开模板 -> 填字段 -> 删旧图 -> 按组平铺插新图 -> 另存为。
    photo_groups: {单元格地址: [图片路径, ...]}，只放图片，不放其他字段。
    """
    wb = load_workbook(template_path)
    ws = wb.active

    fill_fields(ws, fields, packaging_info,
                sample_cartons, sample_numbers, total_cartons)

    summary = insert_photo_groups(ws, photo_groups, image_config)

    wb.save(output_path)
    return summary
